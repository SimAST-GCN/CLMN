import torch 
import torch.nn as nn
import torchvision
import torchvision.transforms as transforms
from torch.utils.data import Dataset, DataLoader,WeightedRandomSampler
from gensim.models.word2vec import Word2Vec
import json
import random
import gensim
import numpy as np
import pandas as pd
import os,time
import warnings
import torch.nn.functional as F
import torch.nn.utils.rnn as rnn_utils
from sklearn import metrics
from sklearn.utils.class_weight import compute_class_weight 
from sklearn.metrics import matthews_corrcoef
import logging
import sys
from openpyxl import Workbook,load_workbook

from data_iter import MADataset
from model import Code_Encoder, CLMN
from eval import CLloss
from transformers import AutoTokenizer, AutoModel

warnings.filterwarnings('ignore')


def set_seed(seed=42):
    random.seed(seed)
    os.environ['PYHTONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
# 设置随机数种子
#set_seed(42)

os.environ["CUDA_VISIBLE_DEVICES"] = "0"
# Device configuration
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
#device = torch.device('cpu')


def run(p,gcnn):
    ot = []

    word2vec = Word2Vec.load('./dataset/'+p+'/node_w2v_512').wv
    MAX_TOKENS = word2vec.syn0.shape[0]

    text_tokenizer = AutoTokenizer.from_pretrained("princeton-nlp/sup-simcse-roberta-large")
    text_encoder = AutoModel.from_pretrained("princeton-nlp/sup-simcse-roberta-large")

    train_path = './dataset/'+p+'/train.pkl'
    test_path = './dataset/'+p+'/test.pkl'
    train_dataset = MADataset(train_path,text_tokenizer)
    test_dataset = MADataset(test_path,text_tokenizer)

    batch_size = 32

    train_loader = DataLoader(dataset = train_dataset,batch_size = batch_size,shuffle = True)
    test_loader = DataLoader(dataset = test_dataset,batch_size = batch_size,shuffle = False)

    epochs = 40    #10 to fine-tuning , 30 for repeat result
    #learning_rate = 0.0002

    pre_train_code_encoder = './result/'+p+'/pre_train_'+str(gcnn)+'.pth'
    code_encoder = torch.load(pre_train_code_encoder)
    model = CLMN(code_encoder,text_encoder).to(device)
    
    l = train_dataset.label
    class_weight = 'balanced'
    classes = np.array([0,1])
    weight = compute_class_weight(class_weight = class_weight,classes = classes, y = l)
    criterion = nn.CrossEntropyLoss(weight = torch.from_numpy(weight).float().cuda())

    #criterion = nn.CrossEntropyLoss()
    #optimizer = torch.optim.Adamax(model.parameters()) #, lr=learning_rate  ,weight_decay=0.00001
    optimizer = torch.optim.Adam(model.parameters(),lr=0.00001,weight_decay=0.00001)#weight_decay=0.00001  #adamax 0.0002
    #optimizer = torch.optim.SGD(model.parameters(), lr=0.01, momentum=0.9, weight_decay=0.00001)
    tm = 0
    total_step = len(train_loader)
    for epoch in range(epochs):
        logging.info("train epoch: "+str(epoch))
        model.train()
        start_time = time.time()
        ts = time.time()
        for _,data in enumerate(train_loader):
            code1,graph1,code2,graph2,input_ids,attention_mask,label = data
            label = label.to(device)
            input_ids = input_ids.to(device)
            attention_mask = attention_mask.to(device)
            graph1 = graph1.to(device).float()
            code1 = code1.to(device).int()
            graph2 = graph2.to(device).float()
            code2 = code2.to(device).int()
            # Forward pass
            outputs = model(code1,graph1,code2,graph2,input_ids,attention_mask)
            loss = criterion(outputs,label)
            # Backward and optimize
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

            if (_+1) % 10 == 0:
                end_time = time.time()
                logging.info('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}, Time: {}' 
                    .format(epoch+1, epochs, _+1, total_step, loss.item(), end_time - start_time))
                start_time = time.time()
        td = time.time()
        tm +=(td-ts)

        model.eval()
        logging.info("testing")
        
        lb = torch.Tensor()
        pr = torch.Tensor()
        total = 0
        correct = 0
        with torch.no_grad():
            for _,data in enumerate(test_loader):
                code1,graph1,code2,graph2,input_ids,attention_mask,label = data
                label = label.to(device)
                input_ids = input_ids.to(device)
                attention_mask = attention_mask.to(device)
                graph1 = graph1.to(device).float()
                code1 = code1.to(device).int()
                graph2 = graph2.to(device).float()
                code2 = code2.to(device).int()
                # Forward pass
                outputs = model(code1,graph1,code2,graph2,input_ids,attention_mask)
                loss = criterion(outputs,label)
                # store result
                __, predicted = torch.max(outputs.data, 1)
                total += label.size(0)
                pr = torch.cat((pr,predicted.cpu()),0)
                lb = torch.cat((lb,label.cpu()),0)
                correct += (predicted == label).sum().item()

                if (_+1) % 10 == 0:
                    end_time = time.time()
                    logging.info('Testing Step [{}/{}], Loss: {:.4f}, Time: {}' 
                        .format(_+1, len(test_loader), loss.item(), end_time - start_time))
                    start_time = time.time()
        zero = 0
        zero_all = 0
        one = 0
        one_all = 0
        for i in range(len(lb)):
            if lb[i]==0:
                zero_all+=1
                if pr[i]==0:zero+=1
            else:
                one_all+=1
                if pr[i]==1:one+=1
        logging.info("Test one acc: {}/{}, zero acc: {}/{}".format(one,one_all,zero,zero_all))
        logging.info("Recall : {}".format(metrics.recall_score(lb,pr)))
        logging.info("F1 : {}".format(metrics.f1_score(lb,pr)))
        logging.info("AUC : {}".format(metrics.roc_auc_score(lb,pr)))
        logging.info("MCC : {}".format(matthews_corrcoef(lb,pr)))
        logging.info("Precision : {}".format(metrics.precision_score(lb,pr)))
        if epoch>=0: ot.append(['','',metrics.recall_score(lb,pr),metrics.f1_score(lb,pr),metrics.roc_auc_score(lb,pr),100*correct/total,matthews_corrcoef(lb,pr),metrics.precision_score(lb,pr),tm])
    return ot

        
if __name__ == '__main__':
    project = sys.argv[1]
    gcnn = sys.argv[2]

    logging.basicConfig(level=logging.INFO,
                        filename='./result/'+project+'/CLMN_'+gcnn+'.log',
                        filemode='a',
                        format='%(asctime)s %(levelname)s %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S')

    out = run(project,int(gcnn))

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'project'
    ws['B1'] = 'model'
    ws['C1'] = 'Recall'
    ws['D1'] = 'F1'
    ws['E1'] = 'AUC'
    ws['F1'] = 'ACCURACY'
    ws['G1'] = 'MCC'
    ws['H1'] = 'Precision'
    ws['I1'] = 'Time'

    file_path = './result/'+project+'/CLMN_'+gcnn+'.xlsx'

    out[0][0]=project
    out[0][1]='CLMN'
    for row in out:
        ws.append(row)
    wb.save(file_path)